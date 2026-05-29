import csv
import io
from datetime import datetime, timezone
from unittest.mock import patch

from openpyxl import load_workbook

from app.models.document import Document
from app.models.journal import Journal, TypeEvenement
from app.models.notification_config import NotificationConfig
from app.models.role import Role
from app.models.source import Source
from app.models.user import User
from app.services.export_service import generer_rapport_excel
from app.services.notification_service import notifier_documents_critiques


def _source(db, **kwargs):
    defaults = dict(
        nom="=cmd|' /C calc'!A0",
        type_serveur="linux",
        protocole="local",
        chemin_distant="/tmp",
    )
    defaults.update(kwargs)
    src = Source(**defaults)
    db.session.add(src)
    db.session.flush()
    return src


def _login(client, db):
    role = Role(nom="journal-role", permissions={"journal.view": True})
    user = User(username="journal-user", role=role)
    user.set_password("secret123")
    db.session.add_all([role, user])
    db.session.commit()
    with client.session_transaction() as session:
        session["_user_id"] = str(user.id)
        session["_fresh"] = True


def test_excel_neutralise_formules_et_assainit_onglets(app, db):
    src = _source(db, nom="=Source/Interne?[x]")
    doc = Document(
        source_id=src.id,
        nom_fichier="+payload.pdf",
        chemin_local="/tmp/payload.pdf",
        date_collecte=datetime.now(timezone.utc),
    )
    db.session.add(doc)
    db.session.commit()

    contenu = generer_rapport_excel()
    wb = load_workbook(io.BytesIO(contenu), data_only=False)

    assert "=Source/Interne?[x]" not in wb.sheetnames
    recap = wb["Recapitulatif"]
    assert recap["A2"].value == "'=Source/Interne?[x]"
    sheet = wb[wb.sheetnames[1]]
    assert sheet["A2"].value == "'+payload.pdf"


def test_export_csv_journaux_neutralise_formules(client, db):
    _login(client, db)
    db.session.add(
        Journal(
            type_evenement=TypeEvenement.ERREUR,
            message="=HYPERLINK(\"http://example.invalid\")",
        )
    )
    db.session.commit()

    response = client.get("/journaux/export.csv")

    assert response.status_code == 200
    rows = list(csv.reader(io.StringIO(response.data.decode("utf-8")), delimiter=";"))
    assert rows[1][4] == "'=HYPERLINK(\"http://example.invalid\")"


def test_notifications_echappent_html(app, db):
    src = _source(db, nom="<script>alert(1)</script>")
    doc = Document(
        source_id=src.id,
        nom_fichier="<img src=x onerror=alert(1)>.pdf",
        chemin_local="/tmp/doc.pdf",
    )
    db.session.add_all([
        doc,
        NotificationConfig(email="ops@example.local", actif=True, notif_critiques=True),
    ])
    db.session.commit()

    with patch("app.services.notification_service.envoyer_email", return_value=True) as mock_email:
        notifier_documents_critiques(src, [doc])

    corps_html = mock_email.call_args.args[2]
    assert "<script>" not in corps_html
    assert "<img" not in corps_html
    assert "&lt;script&gt;" in corps_html
    assert "&lt;img" in corps_html
