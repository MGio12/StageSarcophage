"""
Service d'export des rapports de conformité.

Phase 2 - CDC §8.2 : Export des rapports de conformité.
"""
import io
from datetime import datetime, timezone
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.models.document import Document, StatutDocument
from app.models.source import Source


STATUT_COLORS = {
    StatutDocument.OK: "00AA00",
    StatutDocument.AVERTISSEMENT: "FFA500",
    StatutDocument.CRITIQUE: "FF0000",
    StatutDocument.PURGE: "808080"
}


def generer_rapport_excel(source_id: Optional[int] = None) -> bytes:
    """
    Génère un rapport Excel avec la liste des documents par source.

    Args:
        source_id: Optionnel, filtre sur une source spécifique

    Returns:
        Contenu du fichier XLSX en bytes
    """
    wb = Workbook()
    ws_recap = wb.active
    ws_recap.title = "Recapitulatif"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws_recap.append(["Source", "Documents OK", "Avertissement", "Critique", "Total"])
    for cell in ws_recap[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    sources_query = Source.query.filter(Source.deleted_at.is_(None))
    if source_id:
        sources_query = sources_query.filter(Source.id == source_id)
    sources = sources_query.order_by(Source.nom).all()

    for source in sources:
        docs = Document.query.filter(
            Document.source_id == source.id,
            Document.statut != StatutDocument.PURGE
        ).all()

        ok = sum(1 for d in docs if d.statut == StatutDocument.OK)
        avert = sum(1 for d in docs if d.statut == StatutDocument.AVERTISSEMENT)
        crit = sum(1 for d in docs if d.statut == StatutDocument.CRITIQUE)
        total = len(docs)

        ws_recap.append([source.nom, ok, avert, crit, total])
        for cell in ws_recap[ws_recap.max_row]:
            cell.border = thin_border

        ws_source = wb.create_sheet(title=source.nom[:31])
        ws_source.append(["Fichier", "Statut", "Taille (Ko)", "Date modification", "Date collecte"])
        for cell in ws_source[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for doc in sorted(docs, key=lambda d: d.nom_fichier):
            taille_ko = round(doc.taille_octets / 1024, 1) if doc.taille_octets else ""
            date_mod = doc.date_modification_source.strftime("%d/%m/%Y %H:%M") if doc.date_modification_source else ""
            date_coll = doc.date_collecte.strftime("%d/%m/%Y %H:%M") if doc.date_collecte else ""

            ws_source.append([doc.nom_fichier, doc.statut.value, taille_ko, date_mod, date_coll])

            row = ws_source.max_row
            color = STATUT_COLORS.get(doc.statut, "FFFFFF")
            ws_source.cell(row, 2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws_source.cell(row, 2).font = Font(color="FFFFFF" if doc.statut != StatutDocument.AVERTISSEMENT else "000000")

            for cell in ws_source[row]:
                cell.border = thin_border

        for col in ws_source.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            ws_source.column_dimensions[col_letter].width = min(max_length + 2, 50)

    for col in ws_recap.columns:
        col_letter = col[0].column_letter
        ws_recap.column_dimensions[col_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def generer_rapport_pdf(source_id: Optional[int] = None) -> bytes:
    """
    Génère un rapport PDF avec la liste des documents par source.

    Args:
        source_id: Optionnel, filtre sur une source spécifique

    Returns:
        Contenu du fichier PDF en bytes
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=10
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Heading2"],
        fontSize=12,
        spaceAfter=6
    )

    elements = []

    elements.append(Paragraph("Rapport de conformite - Modes Degrades", title_style))
    elements.append(Paragraph(
        f"Genere le {datetime.now(timezone.utc).strftime('%d/%m/%Y a %H:%M')} UTC",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 10*mm))

    sources_query = Source.query.filter(Source.deleted_at.is_(None))
    if source_id:
        sources_query = sources_query.filter(Source.id == source_id)
    sources = sources_query.order_by(Source.nom).all()

    recap_data = [["Source", "OK", "Avert.", "Crit.", "Total"]]
    for source in sources:
        docs = Document.query.filter(
            Document.source_id == source.id,
            Document.statut != StatutDocument.PURGE
        ).all()
        ok = sum(1 for d in docs if d.statut == StatutDocument.OK)
        avert = sum(1 for d in docs if d.statut == StatutDocument.AVERTISSEMENT)
        crit = sum(1 for d in docs if d.statut == StatutDocument.CRITIQUE)
        recap_data.append([source.nom[:30], str(ok), str(avert), str(crit), str(len(docs))])

    if len(recap_data) > 1:
        elements.append(Paragraph("Recapitulatif par source", subtitle_style))
        recap_table = Table(recap_data, colWidths=[100*mm, 20*mm, 20*mm, 20*mm, 20*mm])
        recap_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066CC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0F0")])
        ]))
        elements.append(recap_table)
        elements.append(Spacer(1, 10*mm))

    for source in sources:
        docs = Document.query.filter(
            Document.source_id == source.id,
            Document.statut != StatutDocument.PURGE
        ).order_by(Document.nom_fichier).all()

        if not docs:
            continue

        elements.append(Paragraph(f"Source : {source.nom}", subtitle_style))

        doc_data = [["Fichier", "Statut", "Taille", "Modification"]]
        for d in docs[:50]:
            taille = f"{d.taille_octets / 1024:.0f} Ko" if d.taille_octets else "-"
            date = d.date_modification_source.strftime("%d/%m/%Y") if d.date_modification_source else "-"
            doc_data.append([
                d.nom_fichier[:40] + ("..." if len(d.nom_fichier) > 40 else ""),
                d.statut.value,
                taille,
                date
            ])

        if len(docs) > 50:
            doc_data.append([f"... et {len(docs) - 50} autres documents", "", "", ""])

        doc_table = Table(doc_data, colWidths=[90*mm, 30*mm, 25*mm, 30*mm])

        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066CC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]

        for i, d in enumerate(docs[:50], start=1):
            if d.statut == StatutDocument.OK:
                bg = colors.HexColor("#D4EDDA")
            elif d.statut == StatutDocument.AVERTISSEMENT:
                bg = colors.HexColor("#FFF3CD")
            elif d.statut == StatutDocument.CRITIQUE:
                bg = colors.HexColor("#F8D7DA")
            else:
                bg = colors.white
            style_commands.append(("BACKGROUND", (1, i), (1, i), bg))

        doc_table.setStyle(TableStyle(style_commands))
        elements.append(doc_table)
        elements.append(Spacer(1, 8*mm))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()
